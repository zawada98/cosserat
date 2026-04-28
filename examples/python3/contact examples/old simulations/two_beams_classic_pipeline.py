# -*- coding: utf-8 -*-
"""
SphereSweptIntersectionMethod – demonstration scene
====================================================

Beam 1 – COMPLETELY FIXED  (acts as a rigid obstacle along +X)
Beam 2 – PARALLEL to Beam 1, CLAMPED at its base, free tip falls under gravity.

Beam 2 starts GAP_Z mm directly above Beam 1.  Gravity bends it downward
until the SSIM component detects contact along their shared length.

Live data collection + 3D plots
---------------------------------
An SSIMPlotter controller records at every step:
  • simulation time  t
  • curvilinear parameter on Beam 1  s1*
  • curvilinear parameter on Beam 2  s2*
  • signed gap  δ  (< 0 ⇒ interpenetration)

At the end of the simulation it generates two 3D plots:
  (1)  δ  vs  (s1 , t)   saved as  delta_vs_s1_time.png
  (2)  δ  vs  (s2 , t)   saved as  delta_vs_s2_time.png

HOW TO STOP THE SIMULATION AND SAVE THE PLOTS
----------------------------------------------
  1. Press  [Q]  inside the runSofa viewer window.
  2. Press  [P]  to save plots without stopping.
  3. Let the simulation run until MAX_STEPS is reached.

Run with:
    runSofa scene_SSIM_TwoBeams.py
"""

import os
import time
import numpy as np
import matplotlib
matplotlib.use("Agg")          # non-interactive backend – safe inside runSofa
import matplotlib.pyplot as plt
from matplotlib import cm
import math
import Sofa
import Sofa.Core

import atexit
import os
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

SCENE_DIR = os.path.dirname(os.path.abspath(__file__))


# ──────────────────────────────────────────────────────────────────────────────
#  Scene parameters
# ──────────────────────────────────────────────────────────────────────────────
BEAM_LENGTH      = 120.0    # [mm]  total length of each beam
NB_SECTIONS      = 6        # number of Cosserat sections per beam
NB_FRAMES        = 20       # number of output Rigid3d frames per beam
RADIUS           = 1.0      # [mm]  cross-section radius
YOUNG_MODULUS    = 3.0e6    # [Pa]
POISSON_RATIO    = 0.49
STIFFNESS        = 1.0e8    # base-clamp stiffness  (Beam 2 clamped end)
ALGORITHM        = "ALGO_1" # "ALGO_1" (segment-seg) or "ALGO_2" (node-seg NR)
DT               = 0.0001     # [s]   time step
MAX_STEPS        = 500      # stop automatically after this many steps
                             # set to 0 to disable auto-stop

# ──────────────────────────────────────────────────────────────────────────────
#  Geometry
# ──────────────────────────────────────────────────────────────────────────────
#
#  Beam 1 – along global +X, centred at Y=0, Z=0  (fully fixed obstacle)
#    base at [0, 0, 0], quaternion [0,0,0,1]
#
#  Beam 2 – along global +Y, base at [L/2, 0, GAP_Z]
#    (so it crosses Beam 1 near the mid-span of both)
#    quaternion for 90° rotation about Z:  [0, 0, sin45°, cos45°]
#    With gravity in -Z the free end bends down and contacts Beam 1.
#
# Initial vertical gap between the two parallel beam axes
GAP_Z = 20.0    # [mm]   must be > 2*RADIUS to start without interpenetration


# ──────────────────────────────────────────────────────────────────────────────
#  Cosserat beam builder
# ──────────────────────────────────────────────────────────────────────────────

def _make_section_params(nb_sections, length):
    sec_len = length / nb_sections
    return (
        [[0., 0., 0.]] * nb_sections,
        [sec_len] * nb_sections,
        [i * sec_len for i in range(nb_sections + 1)],
    )


def _make_frame_params(nb_frames, length):
    fl       = length / nb_frames
    frames   = [[i * fl, 0., 0.,  0., 0., 0., 1.] for i in range(nb_frames + 1)]
    curv_abs = [i * fl for i in range(nb_frames + 1)]
    edge_indices = [[i, i + 1] for i in range(nb_frames)]
    return frames, curv_abs, edge_indices


def _add_collision_model(parent_node, positions, edge_indices, beam_number, nb_frames):

    # Create collision node
    collision_node = parent_node.addChild('CollisionNode')

    # Add topology and collision components
    collision_node.addObject(
        'EdgeSetTopologyContainer',
        name="collisEdgeSet",
        position=positions,
        edges=edge_indices
    )

    collision_node.addObject(
        'EdgeSetTopologyModifier',
        name="colliseEdgeModifier"
    )

    n_pts = nb_frames + 1
    zero_positions = [[0.0, 0.0, 0.0]] * n_pts

    collision_node.addObject('MechanicalObject',
                             name="CollisionDOFs",
                             template='Vec3d',
                             position=zero_positions)

    collision_node.addObject(
        'LineCollisionModel',
        bothSide="1",
        group=str(beam_number),
        proximity=RADIUS
    )
    collision_node.addObject(
        'PointCollisionModel',
        bothSide="1",
        group=str(beam_number),
    )
    collision_node.addObject('RigidMapping',
                             input='@../FramesMO',
                             output='@CollisionDOFs',
                             rigidIndexPerPoint=list(range(n_pts)))


def add_cosserat_beam(parent_node, name, base_pos, base_quat,
                      nb_sections, nb_frames, length, radius,
                      young_modulus, poisson_ratio, stiffness,
                      beam_number,fully_fixed=False):
    """
    Build a Cosserat beam and attach it to *parent_node*.

    Parameters
    ----------
    fully_fixed : bool
        • False  (default) – beam is clamped at its base via a
          RestShapeSpringsForceField and carries a UniformMass so gravity
          acts on it (the tip is free to deform / fall).
        • True  – ALL degrees of freedom are frozen with
          FixedProjectiveConstraint on both the rigid-base MO and the
          cosserat-coordinate MO.  No mass is added.  The beam acts as a
          perfectly rigid, immovable obstacle.

    Returns
    -------
    frame_node : Sofa.Core.Node
        The child node that holds the mapped Rigid3d frames (FramesMO).
    """
    bx, by, bz     = base_pos
    qx, qy, qz, qw = base_quat

    sec_pos, sec_len, curv_in = _make_section_params(nb_sections, length)
    frames, curv_out, edge_indices = _make_frame_params(nb_frames, length)

    beam_node       = parent_node.addChild(name)
    solver_node = beam_node.addChild('solverNode')
    solver_node.addObject('EulerImplicitSolver',
                          rayleighStiffness=0.2, rayleighMass=0.1)
    solver_node.addObject('SparseLDLSolver', name='solver',
                          template='CompressedRowSparseMatrixd')
    #solver_node.addObject('BlockGaussSeidelConstraintSolver', tolerance=1e-5, maxIterations=5e2)
    solver_node.addObject('GenericConstraintCorrection', linearSolver = solver_node.solver.getLinkPath())

    rigid_base_node = solver_node.addChild('rigidBase')

    base_mo = rigid_base_node.addObject(
        'MechanicalObject', template='Rigid3d', name='RigidBaseMO',
        position=[bx, by, bz, qx, qy, qz, qw],
        showObject=True, showObjectScale=3.0)

    if fully_fixed:
        # ── Fix the rigid base completely ──────────────────────────────────
        rigid_base_node.addObject(
            'FixedProjectiveConstraint', name='fixBase',
            indices=[0])
    else:
        # ── Clamp the base with a stiff spring (allows small elastic reaction) ─
        rigid_base_node.addObject(
            'RestShapeSpringsForceField', name='clamp',
            stiffness=stiffness, angularStiffness=stiffness,
            external_points=0, points=0, template='Rigid3d')

    coord_node = solver_node.addChild('cosseratCoordinate')
    coord_mo   = coord_node.addObject(
        'MechanicalObject', template='Vec3d',
        name='cosserat_state', position=sec_pos)
    coord_node.addObject(
        'BeamHookeLawForceField',
        crossSectionShape='circular',
        length=sec_len, radius=radius,
        youngModulus=young_modulus,
        poissonRatio=poisson_ratio)

    if fully_fixed:
        # ── Fix all cosserat strains (zero = straight reference config) ────
        all_section_ids = list(range(nb_sections))
        coord_node.addObject(
            'FixedProjectiveConstraint', name='fixStrains',
            indices=all_section_ids)

    frame_node = solver_node.addChild('mappedFrames')

    frames_mo = frame_node.addObject(
        'MechanicalObject', template='Rigid3d', name='FramesMO',
        position=frames, showObject=True, showObjectScale=2.0)

    if not fully_fixed:
        # Only the free beam needs mass; the fixed beam has none.
        frame_node.addObject('UniformMass', totalMass=0.1)

    frame_node.addObject(
        'DiscreteCosseratMapping', name='cosseratMapping',
        curv_abs_input=curv_in, curv_abs_output=curv_out,
        input1='@../cosseratCoordinate/cosserat_state',
        input2='@../rigidBase/RigidBaseMO',
        output='@FramesMO', debug=False)

    edges_positions = [[x, y, z] for x, y, z, *_ in frames]

    _add_collision_model(parent_node = frame_node,
                         positions = edges_positions,
                         edge_indices=edge_indices,
                         beam_number=beam_number,
                         nb_frames=nb_frames,
                        )

    return frame_node, frames

def add_visual_model(framesNode, frames, rex):
    N = len(frames)
    n_sides = 30
    TWO_PI = 2.0 * math.pi

    def _ring_positions(r):
        return [
            [0.0,
                r * math.cos(TWO_PI * k / n_sides),
             r * math.sin(TWO_PI * k / n_sides),
             ]
            for k in range(n_sides)
        ]

    def _tube_quads(n_frames, n_s):
        quads = []
        for i in range(n_frames - 1):
            for k in range(n_s):
                k1 = (k + 1) % n_s
                v0, v1 = i * n_s + k, i * n_s + k1
                v2, v3 = v0 + n_s, v1 + n_s
                quads.append([v0, v1, v2, v3])
        return quads

    rigid_idx = [i for i in range(N) for _ in range(n_sides)]
    quads = _tube_quads(N, n_sides)
    safe_name = "Tube"

    for r, color_list, suffix in [
        (rex, [0.85, 0.15, 0.15, 1.0], 'outer')
    ]:
        ring_pos = _ring_positions(r) * N
        color_str = " ".join(str(v) for v in color_list)

        vis = framesNode.addChild(f'visual_{suffix}_{safe_name}')

        # Topology (rest positions, connectivity)
        vis.addObject('MeshTopology', name='topo',
                      position=ring_pos, quads=quads)

        # Bug 2 fix: MechanicalObject is the RigidMapping output, not OglModel
        vis.addObject('MechanicalObject', name='visMO',
                      template='Vec3d', position=ring_pos)
        vis.addObject('RigidMapping',
                      input='@../FramesMO',
                      output='@visMO',
                      rigidIndexPerPoint=rigid_idx,
                      globalToLocalCoords=False)

        # OglModel in child node, driven by IdentityMapping from visMO
        ogl = vis.addChild('ogl')
        ogl.addObject('OglModel', name='oglModel',
                      src='@../topo',
                      color=color_str)
        ogl.addObject('IdentityMapping',
                      input='@../visMO',
                      output='@oglModel')


# ── Output directory (same folder as the scene file) ─────────────────────────
SCENE_DIR = os.path.dirname(os.path.abspath(__file__))


class TimingLogger(Sofa.Core.Controller):
    """
    Per-step contact-detection timer with real-time file flushing.

    Every recorded step is written and flushed to disk immediately inside
    _record(), so no data is lost if the scene freezes or crashes.

    Files produced:
        <SCENE_DIR>/<label>_detection_times.txt   (real-time, always)
        <SCENE_DIR>/<label>_detection_times.xlsx  (at exit, if openpyxl available)
    """

    def __init__(self, *args, label="timer", log_interval=50, **kwargs):
        Sofa.Core.Controller.__init__(self, *args, **kwargs)
        self.label           = label
        self.log_interval    = log_interval
        self.detection_times = []       # kept in memory for the xlsx summary
        self._step           = 0

        # Open the txt file immediately and write the header.
        # mode='w' truncates any previous run with the same label.
        self._txt_path = os.path.join(SCENE_DIR,
                                      f"{self.label}_detection_times.txt")
        self._file = open(self._txt_path, "w", buffering=1)  # buffering=1 → line-buffered
        self._file.write("# Contact-detection timing log\n")
        self._file.write(f"# label : {self.label}\n")
        self._file.write("#\n")
        self._file.write("step,detection_time_ms\n")
        # buffering=1 already flushes at every newline, but be explicit:
        self._file.flush()

        # xlsx summary at interpreter exit (best-effort, not critical)
        atexit.register(self._on_exit)

    # ── called by subclasses ──────────────────────────────────────────────────

    def _record(self, elapsed: float):
        """
        Record one step. Writes and flushes the txt file immediately
        so the data is on disk before the next line of Python executes.
        """
        self._step += 1
        elapsed_ms  = elapsed * 1e3
        self.detection_times.append(elapsed_ms)

        # Real-time write + flush — survives a freeze on the very next step
        self._file.write(f"{self._step},{elapsed_ms:.6f}\n")
        self._file.flush()
        os.fsync(self._file.fileno())   # force OS buffer → disk

        if self._step % self.log_interval == 0:
            mean_ms = sum(self.detection_times) / len(self.detection_times)
            print(f"[{self.label}] step {self._step:5d} | "
                  f"mean: {mean_ms:.4f} ms | "
                  f"last: {elapsed_ms:.4f} ms")

    # ── exit hook ─────────────────────────────────────────────────────────────

    def _on_exit(self):
        """
        Close the txt file and append a summary footer.
        Then write the xlsx if openpyxl is available.
        Called by atexit — fires on clean exit AND on most crashes/freezes
        that terminate the Python interpreter.
        """
        if not self.detection_times:
            self._file.close()
            return

        n       = len(self.detection_times)
        mean_ms = sum(self.detection_times) / n
        min_ms  = min(self.detection_times)
        max_ms  = max(self.detection_times)

        # Append summary footer to the txt file
        self._file.write("#\n")
        self._file.write(f"# --- SUMMARY ---\n")
        self._file.write(f"# steps    : {n}\n")
        self._file.write(f"# mean(ms) : {mean_ms:.6f}\n")
        self._file.write(f"# min(ms)  : {min_ms:.6f}\n")
        self._file.write(f"# max(ms)  : {max_ms:.6f}\n")
        self._file.flush()
        self._file.close()
        print(f"[{self.label}] TXT saved → {self._txt_path}")

        # xlsx is written once at the end — best-effort only
        if _HAS_OPENPYXL:
            self._write_xlsx(n, mean_ms, min_ms, max_ms)

    # ── xlsx writer (unchanged from before) ───────────────────────────────────

    def _write_xlsx(self, n, mean_ms, min_ms, max_ms):
        path = os.path.join(SCENE_DIR, f"{self.label}_detection_times.xlsx")
        wb   = openpyxl.Workbook()

        ws_data = wb.active
        ws_data.title = "Detection Times"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(["Step", "Detection Time (ms)"], start=1):
            cell           = ws_data.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, t in enumerate(self.detection_times, start=1):
            ws_data.cell(row=i + 1, column=1, value=i)
            ws_data.cell(row=i + 1, column=2, value=round(t, 6))

        ws_data.column_dimensions["A"].width = 10
        ws_data.column_dimensions["B"].width = 24

        ws_sum = wb.create_sheet("Summary")
        key_font = Font(bold=True)
        key_fill = PatternFill("solid", fgColor="D6E4F0")
        for row_idx, (key, val) in enumerate([
            ("Label",      self.label),
            ("Steps",      n),
            ("Mean (ms)",  round(mean_ms, 6)),
            ("Min (ms)",   round(min_ms,  6)),
            ("Max (ms)",   round(max_ms,  6)),
            ("Total (ms)", round(sum(self.detection_times), 6)),
        ], start=1):
            kc      = ws_sum.cell(row=row_idx, column=1, value=key)
            kc.font = key_font
            kc.fill = key_fill
            ws_sum.cell(row=row_idx, column=2, value=val)

        ws_sum.column_dimensions["A"].width = 16
        ws_sum.column_dimensions["B"].width = 24

        wb.save(path)
        print(f"[{self.label}] XLSX saved → {path}")

# ── Contact-detection timer (classic pipeline) ─────────────────────────────
class CollisionDetectionTimer(TimingLogger):
    """
    Times BruteForceBroadPhase + BVHNarrowPhase + LocalMinDistance by
    hooking CollisionBeginEvent / CollisionEndEvent.
    """
    def __init__(self, *args, **kwargs):
        kwargs.setdefault("label", "classic_pipeline")
        TimingLogger.__init__(self, *args, **kwargs)
        self._t0 = 0.0

    def onCollisionBeginEvent(self, event):
        self._t0 = time.perf_counter()

    def onCollisionEndEvent(self, event):
        self._record(time.perf_counter() - self._t0)

def createScene(root_node: Sofa.Core.Node):
    """
    Scene layout
    ============

    Beam 1 (fixed obstacle)
    -----------------------
    • Lies along the global +X axis at Z = 0.
    • Base at [0, 0, 0], no rotation.
    • ALL DOFs frozen with FixedProjectiveConstraint – it will never move.
    • No mass.

    Beam 2 (cantilever falling under gravity)
    -----------------------------------------
    • Lies along the global +X axis – PARALLEL to Beam 1.
    • Base clamped at [0, 0, GAP_Z] (directly above Beam 1's base), same
      orientation quaternion [0,0,0,1].
    • Clamped at its base by a RestShapeSpringsForceField.
    • Carries a UniformMass; gravity (−Z) bends it downward until it contacts
      Beam 1 along their shared length.
    """

    root_node.gravity = [0., 0., -9810.]   # mm/s² gravity
    root_node.dt      = DT

    root_node.addObject('RequiredPlugin', pluginName=[
        'Sofa.Component.Constraint.Projective',
        'Sofa.Component.LinearSolver.Direct',
        'Sofa.Component.Mass',
        'Sofa.Component.ODESolver.Backward',
        'Sofa.Component.SolidMechanics.Spring',
        'Sofa.Component.Visual',
        'Sofa.GL.Component.Rendering3D',
        'Cosserat',
        'Sofa.Component.StateContainer',
        'Sofa.Component.Setting',
        'Sofa.Component.Constraint.Lagrangian.Correction',
        'Sofa.Component.Mapping.Linear',
        'Sofa.Component.Mapping.NonLinear',
        'Sofa.Component.Topology.Container.Constant',
        "Sofa.Component.Collision.Detection.Intersection",
        "Sofa.Component.Collision.Response.Contact",
        "Sofa.Component.Constraint.Lagrangian.Correction",
        "Sofa.Component.Constraint.Lagrangian.Solver",
        "Sofa.GUI.Component", "Sofa.Component.Collision.Geometry", "Sofa.Component.LinearSolver.Direct",
        "Sofa.Component.Mapping.Linear", "Sofa.Component.MechanicalLoad",
        "Sofa.Component.StateContainer", "Sofa.Component.ODESolver.Backward",
        "Sofa.Component.AnimationLoop",
        "Sofa.Component.Collision.Detection.Algorithm",
        "Sofa.Component.Topology.Container.Dynamic"
    ])

    root_node.addObject('DefaultVisualManagerLoop')
    root_node.addObject('FreeMotionAnimationLoop')
    root_node.addObject('BackgroundSetting', color=[0.05, 0.05, 0.12, 1.0])
    root_node.addObject('CollisionPipeline')
    root_node.addObject('RuleBasedContactManager', responseParams='mu=0.2', response='FrictionContactConstraint')
    root_node.addObject('BruteForceBroadPhase')
    root_node.addObject('BVHNarrowPhase')
    root_node.addObject(CollisionDetectionTimer(name='collisionTimer'))
    root_node.addObject('LocalMinDistance',alarmDistance=2, contactDistance=0.5)

    # Camera positioned to see both beams from a 3/4-angle view
    root_node.addObject('Camera',
                        position=[-40, -80, 120],
                        lookAt=[50, 0, 0])

    root_node.addObject('VisualStyle',
                        displayFlags='showVisualModels showBehaviorModels '
                                     'showCollisionModels '
                                     'hideBoundingCollisionModels '
                                     'hideForceFields '
                                     'hideInteractionForceFields '
                                     'hideWireframe '
                                     'showMechanicalMappings')

    root_node.addObject('BlockGaussSeidelConstraintSolver',tolerance=1e-5, maxIterations=500)

    # ── Beam 1 – horizontal along +X, COMPLETELY FIXED ───────────────────────
    #
    #   Axis:  X ∈ [0, L]   Y = 0   Z = 0
    #   This beam never deforms or translates; it is a rigid obstacle.
    #
    beam1_framesNode, beam1_frames = add_cosserat_beam(
        root_node, 'Beam1_Fixed',
        base_pos   = [0., 0., 0.],
        base_quat  = [0., 0., 0., 1.],
        nb_sections = NB_SECTIONS,
        nb_frames   = NB_FRAMES,
        length      = BEAM_LENGTH,
        radius      = RADIUS,
        young_modulus  = YOUNG_MODULUS,
        poisson_ratio  = POISSON_RATIO,
        stiffness      = STIFFNESS,
        beam_number    = 1,
        fully_fixed    = True,        # ← freeze every DOF
    )


    add_visual_model(beam1_framesNode, beam1_frames, RADIUS)

    # ── Beam 2 – along +X (parallel to Beam 1), CLAMPED at base, free end falls ─
    #
    #   Same orientation as Beam 1 (quaternion [0,0,0,1]).
    #   Base at [0, 0, GAP_Z] so both beams share the same X axis range and
    #   Beam 2 starts GAP_Z mm directly above Beam 1.
    #
    #   Gravity acts in -Z: the free end bends downward until it contacts Beam 1.
    #   The base end (X = 0) is held by the RestShapeSpringsForceField clamp.
    #
    beam2_framesNode, beam2_frames = add_cosserat_beam(
        root_node, 'Beam2_Cantilever',
        base_pos   = [0., 0., GAP_Z],
        base_quat  = [0., 0., 0., 1.],            # same orientation as Beam 1
        nb_sections = NB_SECTIONS,
        nb_frames   = NB_FRAMES,
        length      = BEAM_LENGTH,
        radius      = RADIUS,
        young_modulus  = YOUNG_MODULUS,
        poisson_ratio  = POISSON_RATIO,
        stiffness      = STIFFNESS,
        beam_number    = 2,
        fully_fixed    = False,       # ← free to deform; clamped at base only
    )
    add_visual_model(beam2_framesNode,beam2_frames, RADIUS)

    return root_node