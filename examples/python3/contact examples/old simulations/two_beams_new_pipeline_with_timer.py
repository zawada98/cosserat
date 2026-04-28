# -*- coding: utf-8 -*-
"""
SphereSweptIntersectionMethod – demonstration scene
====================================================

Beam 1 – COMPLETELY FIXED  (acts as a rigid obstacle along +X)
Beam 2 – PARALLEL to Beam 1, CLAMPED at its base, free tip falls under gravity.

Beam 2 starts GAP_Z mm directly above Beam 1.  Gravity bends it downward
until the SSIM component detects contact along their shared length.

Live data collection + 3D plots
---------------------------------
An SSIMPlotter controller records at every step:
  • simulation time  t
  • curvilinear parameter on Beam 1  s1*
  • curvilinear parameter on Beam 2  s2*
  • signed gap  δ  (< 0 ⇒ interpenetration)

At the end of the simulation it generates two 3D plots:
  (1)  δ  vs  (s1 , t)   saved as  delta_vs_s1_time.png
  (2)  δ  vs  (s2 , t)   saved as  delta_vs_s2_time.png

HOW TO STOP THE SIMULATION AND SAVE THE PLOTS
----------------------------------------------
  1. Press  [Q]  inside the runSofa viewer window.
  2. Press  [P]  to save plots without stopping.
  3. Let the simulation run until MAX_STEPS is reached.

Run with:
    runSofa scene_SSIM_TwoBeams.py
"""

import os
import time
import numpy as np
import matplotlib
matplotlib.use("Agg")          # non-interactive backend – safe inside runSofa
import matplotlib.pyplot as plt
from matplotlib import cm
import math
import Sofa
import Sofa.Core

import atexit
import os
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

SCENE_DIR = os.path.dirname(os.path.abspath(__file__))


# ──────────────────────────────────────────────────────────────────────────────
#  Scene parameters
# ──────────────────────────────────────────────────────────────────────────────
BEAM_LENGTH      = 120.0    # [mm]  total length of each beam
NB_SECTIONS      = 6        # number of Cosserat sections per beam
NB_FRAMES        = 20       # number of output Rigid3d frames per beam
RADIUS           = 1.0      # [mm]  cross-section radius
YOUNG_MODULUS    = 3.0e6    # [Pa]
POISSON_RATIO    = 0.49
STIFFNESS        = 1.0e8    # base-clamp stiffness  (Beam 2 clamped end)
ALGORITHM        = "ALGO_1" # "ALGO_1" (segment-seg) or "ALGO_2" (node-seg NR)
DT               = 0.0001     # [s]   time step
MAX_STEPS        = 500      # stop automatically after this many steps
MAX_K = NB_FRAMES
ALARM_DISTANCE = 2.0 * RADIUS

# ──────────────────────────────────────────────────────────────────────────────
#  Geometry
# ──────────────────────────────────────────────────────────────────────────────
#
#  Beam 1 – along global +X, centred at Y=0, Z=0  (fully fixed obstacle)
#    base at [0, 0, 0], quaternion [0,0,0,1]
#
#  Beam 2 – along global +Y, base at [L/2, 0, GAP_Z]
#    (so it crosses Beam 1 near the mid-span of both)
#    quaternion for 90° rotation about Z:  [0, 0, sin45°, cos45°]
#    With gravity in -Z the free end bends down and contacts Beam 1.
#
# Initial vertical gap between the two parallel beam axes
GAP_Z = 20.0    # [mm]   must be > 2*RADIUS to start without interpenetration

def extract_contact_points(parent_node, contactMO, MAX_K):
    even_indices = list(range(0, 2 * MAX_K, 2))
    odd_indices  = list(range(1, 2 * MAX_K, 2))

    node_A = parent_node.addChild('Pc_A')
    contactMO_A = node_A.addObject(
        'MechanicalObject', template='Vec3d',
        name='contactMO_A',
        position=[[0., 0., 0.]] * MAX_K)
    node_A.addObject(
        'SubsetMapping', template='Vec3d,Vec3d',
        input=contactMO.getLinkPath(),
        output='@contactMO_A',
        indices=even_indices,
        handleTopologyChange=False)

    node_B = parent_node.addChild('Pc_B')
    contactMO_B = node_B.addObject(
        'MechanicalObject', template='Vec3d',
        name='contactMO_B',
        position=[[0., 0., 0.]] * MAX_K)
    node_B.addObject(
        'SubsetMapping', template='Vec3d,Vec3d',
        input=contactMO.getLinkPath(),
        output='@contactMO_B',
        indices=odd_indices,
        handleTopologyChange=False)

    return contactMO_A, contactMO_B
# ──────────────────────────────────────────────────────────────────────────────
#  Cosserat beam builder
# ──────────────────────────────────────────────────────────────────────────────

def _make_section_params(nb_sections, length):
    sec_len = length / nb_sections
    return (
        [[0., 0., 0.]] * nb_sections,
        [sec_len] * nb_sections,
        [i * sec_len for i in range(nb_sections + 1)],
    )


def _make_frame_params(nb_frames, length):
    fl       = length / nb_frames
    frames   = [[i * fl, 0., 0.,  0., 0., 0., 1.] for i in range(nb_frames + 1)]
    curv_abs = [i * fl for i in range(nb_frames + 1)]
    edge_indices = [[i, i + 1] for i in range(nb_frames)]
    return frames, curv_abs, edge_indices

def add_cosserat_beam(parent_node, name, base_pos, base_quat,
                      nb_sections, nb_frames, length, radius,
                      young_modulus, poisson_ratio, stiffness,
                      beam_number,fully_fixed=False):
    """
    Build a Cosserat beam and attach it to *parent_node*.

    Parameters
    ----------
    fully_fixed : bool
        • False  (default) – beam is clamped at its base via a
          RestShapeSpringsForceField and carries a UniformMass so gravity
          acts on it (the tip is free to deform / fall).
        • True  – ALL degrees of freedom are frozen with
          FixedProjectiveConstraint on both the rigid-base MO and the
          cosserat-coordinate MO.  No mass is added.  The beam acts as a
          perfectly rigid, immovable obstacle.

    Returns
    -------
    frame_node : Sofa.Core.Node
        The child node that holds the mapped Rigid3d frames (FramesMO).
    """
    bx, by, bz     = base_pos
    qx, qy, qz, qw = base_quat

    sec_pos, sec_len, curv_in = _make_section_params(nb_sections, length)
    frames, curv_out, edge_indices = _make_frame_params(nb_frames, length)

    beam_node       = parent_node.addChild(name)
    solver_node = beam_node.addChild('solverNode')
    solver_node.addObject('EulerImplicitSolver',
                          rayleighStiffness=0.2, rayleighMass=0.1)
    solver_node.addObject('SparseLDLSolver', name='solver',
                          template='CompressedRowSparseMatrixd')
    solver_node.addObject('GenericConstraintCorrection', linearSolver = solver_node.solver.getLinkPath())

    rigid_base_node = solver_node.addChild('rigidBase')

    base_mo = rigid_base_node.addObject(
        'MechanicalObject', template='Rigid3d', name='RigidBaseMO',
        position=[bx, by, bz, qx, qy, qz, qw],
        showObject=True, showObjectScale=3.0)

    if fully_fixed:
        rigid_base_node.addObject(
            'FixedProjectiveConstraint', name='fixBase',
            indices=[0])
    else:
        rigid_base_node.addObject(
            'RestShapeSpringsForceField', name='clamp',
            stiffness=stiffness, angularStiffness=stiffness,
            external_points=0, points=0, template='Rigid3d')

    coord_node = solver_node.addChild('cosseratCoordinate')
    coord_mo   = coord_node.addObject(
        'MechanicalObject', template='Vec3d',
        name='cosserat_state', position=sec_pos)
    coord_node.addObject(
        'BeamHookeLawForceField',
        crossSectionShape='circular',
        length=sec_len, radius=radius,
        youngModulus=young_modulus,
        poissonRatio=poisson_ratio)

    if fully_fixed:
        all_section_ids = list(range(nb_sections))
        coord_node.addObject(
            'FixedProjectiveConstraint', name='fixStrains',
            indices=all_section_ids)

    frame_node = solver_node.addChild('mappedFrames')

    frames_mo = frame_node.addObject(
        'MechanicalObject', template='Rigid3d', name='FramesMO',
        position=frames, showObject=True, showObjectScale=2.0)

    if not fully_fixed:
        frame_node.addObject('UniformMass', totalMass=0.1)

    frame_node.addObject(
        'DiscreteCosseratMapping', name='cosseratMapping',
        curv_abs_input=curv_in, curv_abs_output=curv_out,
        input1='@../cosseratCoordinate/cosserat_state',
        input2='@../rigidBase/RigidBaseMO',
        output='@FramesMO', debug=False)

    edges_positions = [[x, y, z] for x, y, z, *_ in frames]

    return frame_node, frames

def add_visual_model(framesNode, frames, rex):
    N = len(frames)
    n_sides = 30
    TWO_PI = 2.0 * math.pi

    def _ring_positions(r):
        return [
            [0.0,
                r * math.cos(TWO_PI * k / n_sides),
             r * math.sin(TWO_PI * k / n_sides),
             ]
            for k in range(n_sides)
        ]

    def _tube_quads(n_frames, n_s):
        quads = []
        for i in range(n_frames - 1):
            for k in range(n_s):
                k1 = (k + 1) % n_s
                v0, v1 = i * n_s + k, i * n_s + k1
                v2, v3 = v0 + n_s, v1 + n_s
                quads.append([v0, v1, v2, v3])
        return quads

    rigid_idx = [i for i in range(N) for _ in range(n_sides)]
    quads = _tube_quads(N, n_sides)
    safe_name = "Tube"

    for r, color_list, suffix in [
        (rex, [0.85, 0.15, 0.15, 1.0], 'outer')
    ]:
        ring_pos = _ring_positions(r) * N
        color_str = " ".join(str(v) for v in color_list)

        vis = framesNode.addChild(f'visual_{suffix}_{safe_name}')

        vis.addObject('MeshTopology', name='topo',
                      position=ring_pos, quads=quads)

        vis.addObject('MechanicalObject', name='visMO',
                      template='Vec3d', position=ring_pos)
        vis.addObject('RigidMapping',
                      input='@../FramesMO',
                      output='@visMO',
                      rigidIndexPerPoint=rigid_idx,
                      globalToLocalCoords=False)

        ogl = vis.addChild('ogl')
        ogl.addObject('OglModel', name='oglModel',
                      src='@../topo',
                      color=color_str)
        ogl.addObject('IdentityMapping',
                      input='@../visMO',
                      output='@oglModel')


# ── Output directory (same folder as the scene file) ─────────────────────────
SCENE_DIR = os.path.dirname(os.path.abspath(__file__))


# ──────────────────────────────────────────────────────────────────────────────
#  Timing infrastructure  (mirrors classic pipeline's TimingLogger exactly)
# ──────────────────────────────────────────────────────────────────────────────

class TimingLogger(Sofa.Core.Controller):
    """
    Per-step contact-detection timer with real-time file flushing.

    Every recorded step is written and flushed to disk immediately inside
    _record(), so no data is lost if the scene freezes or crashes.

    Files produced:
        <SCENE_DIR>/<label>_detection_times.txt   (real-time, always)
        <SCENE_DIR>/<label>_detection_times.xlsx  (at exit, if openpyxl available)
    """

    def __init__(self, *args, label="timer", log_interval=50, **kwargs):
        Sofa.Core.Controller.__init__(self, *args, **kwargs)
        self.label           = label
        self.log_interval    = log_interval
        self.detection_times = []
        self._step           = 0

        self._txt_path = os.path.join(SCENE_DIR,
                                      f"{self.label}_detection_times.txt")
        self._file = open(self._txt_path, "w", buffering=1)
        self._file.write("# Contact-detection timing log\n")
        self._file.write(f"# label : {self.label}\n")
        self._file.write("#\n")
        self._file.write("step,detection_time_ms\n")
        self._file.flush()

        atexit.register(self._on_exit)

    def _record(self, elapsed: float):
        """
        Record one step. Writes and flushes the txt file immediately.
        """
        self._step += 1
        elapsed_ms  = elapsed * 1e3
        self.detection_times.append(elapsed_ms)

        self._file.write(f"{self._step},{elapsed_ms:.6f}\n")
        self._file.flush()
        os.fsync(self._file.fileno())

        if self._step % self.log_interval == 0:
            mean_ms = sum(self.detection_times) / len(self.detection_times)
            print(f"[{self.label}] step {self._step:5d} | "
                  f"mean: {mean_ms:.4f} ms | "
                  f"last: {elapsed_ms:.4f} ms")

    def _on_exit(self):
        if not self.detection_times:
            self._file.close()
            return

        n       = len(self.detection_times)
        mean_ms = sum(self.detection_times) / n
        min_ms  = min(self.detection_times)
        max_ms  = max(self.detection_times)

        self._file.write("#\n")
        self._file.write(f"# --- SUMMARY ---\n")
        self._file.write(f"# steps    : {n}\n")
        self._file.write(f"# mean(ms) : {mean_ms:.6f}\n")
        self._file.write(f"# min(ms)  : {min_ms:.6f}\n")
        self._file.write(f"# max(ms)  : {max_ms:.6f}\n")
        self._file.flush()
        self._file.close()
        print(f"[{self.label}] TXT saved → {self._txt_path}")

        if _HAS_OPENPYXL:
            self._write_xlsx(n, mean_ms, min_ms, max_ms)

    def _write_xlsx(self, n, mean_ms, min_ms, max_ms):
        path = os.path.join(SCENE_DIR, f"{self.label}_detection_times.xlsx")
        wb   = openpyxl.Workbook()

        ws_data = wb.active
        ws_data.title = "Detection Times"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(["Step", "Detection Time (ms)"], start=1):
            cell           = ws_data.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, t in enumerate(self.detection_times, start=1):
            ws_data.cell(row=i + 1, column=1, value=i)
            ws_data.cell(row=i + 1, column=2, value=round(t, 6))

        ws_data.column_dimensions["A"].width = 10
        ws_data.column_dimensions["B"].width = 24

        ws_sum = wb.create_sheet("Summary")
        key_font = Font(bold=True)
        key_fill = PatternFill("solid", fgColor="D6E4F0")
        for row_idx, (key, val) in enumerate([
            ("Label",      self.label),
            ("Steps",      n),
            ("Mean (ms)",  round(mean_ms, 6)),
            ("Min (ms)",   round(min_ms,  6)),
            ("Max (ms)",   round(max_ms,  6)),
            ("Total (ms)", round(sum(self.detection_times), 6)),
        ], start=1):
            kc      = ws_sum.cell(row=row_idx, column=1, value=key)
            kc.font = key_font
            kc.fill = key_fill
            ws_sum.cell(row=row_idx, column=2, value=val)

        ws_sum.column_dimensions["A"].width = 16
        ws_sum.column_dimensions["B"].width = 24

        wb.save(path)
        print(f"[{self.label}] XLSX saved → {path}")


# ── Contact-detection timer (SSIM pipeline) ───────────────────────────────────
class SSIMDetectionTimer(TimingLogger):
    """
    Times the SphereSweptIntersectionMethod detection phase by forcing a
    fresh SSIM update at the start of each animate step and measuring how
    long it takes to read back the distances output.

    Timing scope
    ------------
    The SSIM is a DataEngine — it runs lazily when its outputs are first
    accessed.  This controller calls ssim.update() explicitly at
    onAnimateBeginEvent and measures only that call, giving a direct
    measure of the SSIM detection cost isolated from the rest of the step.

    This is the exact equivalent of CollisionBeginEvent→CollisionEndEvent
    in the classic pipeline (BruteForceBroadPhase + BVHNarrowPhase +
    LocalMinDistance).

    Output files (same format as classic pipeline):
        ssim_pipeline_detection_times.txt
        ssim_pipeline_detection_times.xlsx
    """

    def __init__(self, ssim_component, *args, **kwargs):
        kwargs.setdefault("label", "ssim_pipeline")
        TimingLogger.__init__(self, *args, **kwargs)
        self._ssim = ssim_component

    def onAnimateBeginEvent(self, event):
        t0 = time.perf_counter()
        # Reading an output Data field forces the DataEngine to recompute.
        # This is equivalent to calling update() — SofaPython3 does not
        # expose update() directly on DataEngine subclasses.
        _ = self._ssim.distances.value
        self._record(time.perf_counter() - t0)


def createScene(root_node: Sofa.Core.Node):
    """
    Scene layout
    ============

    Beam 1 (fixed obstacle)
    -----------------------
    • Lies along the global +X axis at Z = 0.
    • Base at [0, 0, 0], no rotation.
    • ALL DOFs frozen with FixedProjectiveConstraint – it will never move.
    • No mass.

    Beam 2 (cantilever falling under gravity)
    -----------------------------------------
    • Lies along the global +X axis – PARALLEL to Beam 1.
    • Base clamped at [0, 0, GAP_Z] (directly above Beam 1's base), same
      orientation quaternion [0,0,0,1].
    • Clamped at its base by a RestShapeSpringsForceField.
    • Carries a UniformMass; gravity (−Z) bends it downward until it contacts
      Beam 1 along their shared length.
    """

    root_node.gravity = [0., 0., -9810.]   # mm/s² gravity
    root_node.dt      = DT

    root_node.addObject('RequiredPlugin', pluginName=[
        'Sofa.Component.Constraint.Projective',
        'Sofa.Component.LinearSolver.Direct',
        'Sofa.Component.Mass',
        'Sofa.Component.ODESolver.Backward',
        'Sofa.Component.SolidMechanics.Spring',
        'Sofa.Component.Visual',
        'Sofa.GL.Component.Rendering3D',
        'Cosserat',
        'Sofa.Component.StateContainer',
        'Sofa.Component.Setting',
        'Sofa.Component.Constraint.Lagrangian.Correction',
        'Sofa.Component.Mapping.Linear',
        'Sofa.Component.Mapping.NonLinear',
        'Sofa.Component.Topology.Container.Constant',
        "Sofa.Component.Collision.Detection.Intersection",
        "Sofa.Component.Collision.Response.Contact",
        "Sofa.Component.Constraint.Lagrangian.Correction",
        "Sofa.Component.Constraint.Lagrangian.Solver",
        "Sofa.GUI.Component", "Sofa.Component.Collision.Geometry", "Sofa.Component.LinearSolver.Direct",
        "Sofa.Component.Mapping.Linear", "Sofa.Component.MechanicalLoad",
        "Sofa.Component.StateContainer", "Sofa.Component.ODESolver.Backward",
        "Sofa.Component.AnimationLoop",
        "Sofa.Component.Collision.Detection.Algorithm",
        "Sofa.Component.Topology.Container.Dynamic",
        'Sofa.Component.Constraint.Lagrangian.Model'
    ])

    root_node.addObject('DefaultVisualManagerLoop')
    root_node.addObject('FreeMotionAnimationLoop')
    root_node.addObject('BackgroundSetting', color=[0.05, 0.05, 0.12, 1.0])

    root_node.addObject('Camera',
                        position=[-40, -80, 120],
                        lookAt=[50, 0, 0])

    root_node.addObject('VisualStyle',
                        displayFlags='showVisualModels showBehaviorModels '
                                     'showCollisionModels '
                                     'hideBoundingCollisionModels '
                                     'hideForceFields '
                                     'hideInteractionForceFields '
                                     'hideWireframe '
                                     'showMechanicalMappings')

    root_node.addObject('BlockGaussSeidelConstraintSolver',tolerance=1e-5, maxIterations=500)

    # ── Beam 1 – horizontal along +X, COMPLETELY FIXED ───────────────────────
    beam1_framesNode, beam1_frames = add_cosserat_beam(
        root_node, 'Beam1_Fixed',
        base_pos   = [0., 0., 0.],
        base_quat  = [0., 0., 0., 1.],
        nb_sections = NB_SECTIONS,
        nb_frames   = NB_FRAMES,
        length      = BEAM_LENGTH,
        radius      = RADIUS,
        young_modulus  = YOUNG_MODULUS,
        poisson_ratio  = POISSON_RATIO,
        stiffness      = STIFFNESS,
        beam_number    = 1,
        fully_fixed    = True,
    )

    add_visual_model(beam1_framesNode, beam1_frames, RADIUS)

    # ── Beam 2 – along +X (parallel to Beam 1), CLAMPED at base, free end falls ─
    beam2_framesNode, beam2_frames = add_cosserat_beam(
        root_node, 'Beam2_Cantilever',
        base_pos   = [0., 0., GAP_Z],
        base_quat  = [0., 0., 0., 1.],
        nb_sections = NB_SECTIONS,
        nb_frames   = NB_FRAMES,
        length      = BEAM_LENGTH,
        radius      = RADIUS,
        young_modulus  = YOUNG_MODULUS,
        poisson_ratio  = POISSON_RATIO,
        stiffness      = STIFFNESS,
        beam_number    = 2,
        fully_fixed    = False,
    )
    add_visual_model(beam2_framesNode, beam2_frames, RADIUS)

    intersection_node = root_node.addChild('IntersectionNode')

    beam1_MO = beam1_framesNode.FramesMO
    beam2_MO = beam2_framesNode.FramesMO

    ssim = intersection_node.addObject(
        'SphereSweptIntersectionMethod',
        name='ssim',
        beam1Frames=beam1_MO.getLinkPath() + '.position',
        beam2Frames=beam2_MO.getLinkPath() + '.position',
        radius1=RADIUS,
        radius2=RADIUS,
        algorithmType=ALGORITHM,
    )

    # ── SSIM detection timer (mirrors CollisionDetectionTimer in classic pipeline) ──
    # SSIMDetectionTimer.onAnimateBeginEvent calls ssim.update() and times it.
    # This isolates the SSIM detection cost from constraint solving and mechanics,
    # making it directly comparable to CollisionBeginEvent→CollisionEndEvent timing
    # in the classic pipeline.
    root_node.addObject(SSIMDetectionTimer(
        ssim_component=ssim,
        name='ssimTimer',
        log_interval=50,
    ))

    contact_output = beam1_framesNode.addChild('contactOutput')
    beam2_framesNode.addChild(contact_output)

    contactMO = contact_output.addObject(
        'MechanicalObject', template='Vec3d',
        name='contactMO_gap',
        position=[[0., 0., 0.]] * 2*MAX_K)

    contact_output.addObject(
        'BeamContactMapping',
        name='bcm',
        input1=beam1_MO.getLinkPath(),
        input2=beam2_MO.getLinkPath(),
        output=contactMO.getLinkPath(),
        radius1=RADIUS,
        radius2=RADIUS,
        isAlgo2=ALGORITHM == 'ALGO_2',
        mappingMode='contactPoints',
        contactSectionIds=ssim.getLinkPath() + '.contactSectionIds',
        curvilinearParams=ssim.getLinkPath() + '.curvilinearParams',
    )

    contact_output.addObject(
        'UnilateralLagrangianConstraint',
        template='Vec3d',
        name='ulc',
        object1=contactMO.getLinkPath(),
        object2=contactMO.getLinkPath(),
        )

    cf = contact_output.addObject(
        'ContactFeeder',
        name='feeder',
        surfacePoints1=ssim.getLinkPath() + '.surfacePoints1',
        surfacePoints2=ssim.getLinkPath() + '.surfacePoints2',
        distances=ssim.getLinkPath() + '.distances',
        centerlinePoints1=ssim.getLinkPath() + '.centerlinePoints1',
        centerlinePoints2=ssim.getLinkPath() + '.centerlinePoints2',
        constraint='@ulc',
        alarmDistance=ALARM_DISTANCE,
        mu=0.0,
    )
    return root_node